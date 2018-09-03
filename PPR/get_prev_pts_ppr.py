from lxml import html
import requests
import re
from openpyxl import Workbook
from openpyxl import load_workbook


filename = "C:\\Users\\Avery\\Desktop\\Sports Analytics\\FantasyFootball2018\\PPR\\PlayerData_PPR.xlsx"
wb = load_workbook(filename)

filename_data = "C:\\Users\\Avery\\Desktop\\Sports Analytics\\FantasyFootball2018\\PrevYearPts.xlsx"
wb_data = load_workbook(filename_data)
ws_data = wb_data["Worksheet"]


print("Matching Players to their Stats from Last Year")


def match_player_by_pos(ws):
	ws['C1'] = "Prev_Yr_Pts"
	for j in range(2, ws.max_row + 1):
		player = ws['A' + str(j)].value
		player = player.lower()
		if player in player_name:
			ws['C' + str(j)] = float(player_pts[player_name.index(player)])
		else:
			ws['C' + str(j)] = 0


ws_qb = wb["QB"]
ws_rb = wb["RB"]
ws_wr = wb["WR"]
ws_te = wb["TE"]


player_name = []
player_pts = []
for i in range(3, 502):
	player = ws_data['B' + str(i)].value
	player = player.replace('*', '')
	player = player.replace('+', '')
	player = player.lower()
	player_name.append(player)
	player_pts.append(ws_data['Y' + str(i)].value)
#print(player_name)
#print(player_pts)

print("Matching QBs to their Stats Last Year")
match_player_by_pos(ws_qb)

print("Matching RBs to their Stats Last Year")
match_player_by_pos(ws_rb)

print("Matching WRs to their Stats Last Year")
match_player_by_pos(ws_wr)

print("Matching TEs to their Stats Last Year")
match_player_by_pos(ws_te)


wb.save(filename)
print("DONE!")