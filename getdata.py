from lxml import html
import requests
import re
from openpyxl import Workbook


filename = "C:\\Users\\Avery\\Desktop\\Sports Analytics\\FantasyFootball2018\\PlayerData.xlsx"
wb = Workbook()


def get_espn_proj(position, slot_cat_id, num_players):
	ws = wb.create_sheet(position)
	ws['A1'] = "Player"
	ws['B1'] = "ESPN_Proj"

	player_data = [['' for x in range(2)] for y in range(num_players)]

	for player_index in range(0, num_players, 40):
		page = requests.get("http://games.espn.com/ffl/tools/projections?slotCategoryId=" + str(slot_cat_id) + "&leagueId=728465&startIndex=" + str(player_index))
		tree = html.fromstring(page.content)
		
		if (num_players - player_index < 40):
			stop_index = num_players - player_index
		else:
			stop_index = 40

		data_long = tree.xpath("//td//text()")
		data_long = data_long[(data_long.index('PTS')+2):]
		if (position == "D_ST"):
			player_name = re.compile("\w+\sD\/ST")
		else:
			player_name = re.compile("[a-zA-z\.']+\s\w+")
		players = list(filter(player_name.match, data_long))
		players = players[0:stop_index]

		data_short = tree.xpath("//td/text()")
		espn_proj = []
		if (position == "D_ST"):
			cat_start = 16
			cat_spaces = 11
		elif (position == "K"):
			cat_start = 14
			cat_spaces = 9
		else:
			cat_spaces = 14
			cat_start = 20
		for i in range(cat_start, len(data_short), cat_spaces):
			espn_proj.append(data_short[i])
			
		print(players)
		print(espn_proj)
			
		for i in range(0, stop_index):
			print("i: " + str(i))
			player_data[i + player_index][0] = players[i]
			ws['A' + str(i + player_index + 2)] = players[i]
			player_data[i + player_index][1] = espn_proj[i]
			ws['B' + str(i + player_index + 2)] = float(espn_proj[i])

	#print(players)
	#print(espn_proj)
	print(player_data)
	wb.save(filename)
	
	
get_espn_proj("QB", 0, 120)
get_espn_proj("RB", 2, 120)
get_espn_proj("WR", 4, 120)
get_espn_proj("TE", 6, 120)
get_espn_proj("D_ST", 16, 32)
get_espn_proj("K", 17, 49)

